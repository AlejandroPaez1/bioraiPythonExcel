import csv
import os
from tkinter import Tk, filedialog, messagebox
from datetime import datetime, timedelta
import openpyxl
from collections import defaultdict
import re

def determinar_formato(checada_str):
    if '/' in checada_str:
        if len(checada_str) == 16:
            return "%d/%m/%Y %H:%M"
        elif len(checada_str) == 19:
            return "%d/%m/%Y %H:%M:%S"
    elif '-' in checada_str:
        if len(checada_str) == 16:
            return "%d-%m-%Y %H:%M"
        elif len(checada_str) == 19:
            return "%d-%m-%Y %H:%M:%S"
    raise ValueError("Formato de fecha y hora no reconocido")

def obtener_nombre_archivo(nombre_base, extension):
    contador = 1
    while True:
        nombre_archivo = f"{nombre_base}_({contador}).{extension}"
        if not os.path.exists(nombre_archivo):
            return nombre_archivo
        contador += 1

def buscar_no_checadores(archivo_checadores):
    with open(archivo_checadores, 'r', newline='', encoding='utf-8') as file:
        csv_reader = csv.DictReader(file)
        data_checadores = list(csv_reader)

    checadas_por_pin = defaultdict(list)
    for empleado in data_checadores:
        pin = empleado["PIN"]
        nombre = empleado["Nombre de empleado"]
        dispositivo = empleado["Dispositivo"]
        checada_str = empleado["Checada"]
        formato = determinar_formato(checada_str)
        checada_dt = datetime.strptime(checada_str, formato)
        checadas_por_pin[pin].append((checada_dt, nombre, dispositivo))

    todas_las_fechas = [checada.date() for checadas in checadas_por_pin.values() for checada, _, _ in checadas]
    primer_dia_global = min(todas_las_fechas)
    ultimo_dia_global = max(todas_las_fechas)

    todas_las_fechas_posibles = [primer_dia_global + timedelta(days=i) for i in range((ultimo_dia_global - primer_dia_global).days + 1)]
    dias_no_checados_por_pin = defaultdict(list)

    for pin, checadas in checadas_por_pin.items():
        dias_checados = set(checada.date() for checada, _, _ in checadas)
        for fecha in todas_las_fechas_posibles:
            if fecha not in dias_checados:
                dias_no_checados_por_pin[pin].append(fecha)

    return dias_no_checados_por_pin, data_checadores

def extraer_numeros_checada(celda):
    patron_horaminuto = r'\b\d{2}:\d{2}\b'
    horas_minutos_encontrados = re.findall(patron_horaminuto, celda)
    if horas_minutos_encontrados:
        return horas_minutos_encontrados
    return None

def obtener_numeros_checada_por_pin_y_fecha(pin, fecha, data_checadores):
    numeros_checada = []
    fecha_str = fecha.strftime("%d-%m-%Y")
    for empleado in data_checadores:
        if empleado['PIN'] == pin and empleado['Checada'].startswith(fecha_str):
            numeros_checada.extend(extraer_numeros_checada(empleado['Checada']))
    return numeros_checada

def buscar_nombre_por_pin(pin, data_checadores):
    for empleado in data_checadores:
        if empleado['PIN'] == pin:
            return empleado['Nombre de empleado']
    return None

def buscar_dispositivo_por_pin(pin, data_checadores):
    for empleado in data_checadores:
        if empleado['PIN'] == pin:
            return empleado['Dispositivo']
    return None

def main():
    root = Tk()
    root.withdraw()

    messagebox.showinfo("Rango de fechas", "Por favor, asegúrese de seleccionar un archivo CSV que contenga registros de checadas de hasta una semana atrás hasta la fecha actual.")

    archivo_checadores = filedialog.askopenfilename(title="Seleccione el archivo CSV", filetypes=[("Archivos CSV", "*.csv")])

    try:
        dias_no_checados_por_empleado, data_checadores = buscar_no_checadores(archivo_checadores)
        
        wb = openpyxl.Workbook()
        ws = wb.active

        ws['A1'] = "PIN"
        ws['B1'] = "Nombre"
        ws['C1'] = "Dispositivo"

        fechas = sorted(set(fecha for dias_no_checados in dias_no_checados_por_empleado.values() for fecha in dias_no_checados))

        for idx, fecha in enumerate(fechas, start=1):
            ws.cell(row=1, column=3+idx).value = fecha.strftime("%d/%m/%Y")

        row = 2
        for pin, dias_no_checados in dias_no_checados_por_empleado.items():
            if dias_no_checados:
                ws.cell(row=row, column=1).value = int(pin)
                ws.cell(row=row, column=2).value = buscar_nombre_por_pin(pin, data_checadores)
                ws.cell(row=row, column=3).value = buscar_dispositivo_por_pin(pin, data_checadores)

                for idx, fecha in enumerate(fechas, start=1):
                    if fecha in dias_no_checados:
                        ws.cell(row=row, column=3+idx).value = "F"
                    else:
                        numeros_checada = obtener_numeros_checada_por_pin_y_fecha(pin, fecha, data_checadores)
                        if numeros_checada:
                            ws.cell(row=row, column=3+idx).value = ', '.join(numeros_checada)
                row += 1

        nombre_base = "resultadoFaltas"
        extension = "xlsx"
        nombre_archivo = obtener_nombre_archivo(nombre_base, extension)
        wb.save(nombre_archivo)

        messagebox.showinfo("Archivo creado", f"Se ha creado el archivo '{nombre_archivo}' con éxito.")
    except Exception as e:
        print("Error", f"Se ha producido un error: {str(e)}")
        messagebox.showerror("Error", f"Se ha producido un error: {str(e)}")

if __name__ == "__main__":
    main()
