import csv
import os
from tkinter import Tk, filedialog,messagebox
from datetime import datetime, timedelta
import openpyxl
from collections import defaultdict
try:
    def determinar_formato(checada_str):
        if '/' in checada_str:
            if len(checada_str) == 16:
                return "%d/%m/%Y %H:%M"
            elif len(checada_str) == 19:
                return "%d/%m/%Y %H:%M:%S"
        elif '-' in checada_str:
            if len(checada_str) == 16:
                return "%d-%m-%Y %H:%M"
            elif len(checada_str) == 19:
                return "%d-%m-%Y %H:%M:%S"
        else:
            messagebox.showinfo("Formato de fecha y hora no reconocido")
            # print("Formato de fecha y hora no reconocido")
            raise ValueError("Formato de fecha y hora no reconocido")
        
    def obtener_nombre_archivo(nombre_base, extension):
        contador = 1
        while True:
            nombre_archivo = f"{nombre_base}_({contador}).{extension}"
            if not os.path.exists(nombre_archivo):
                return nombre_archivo
            contador += 1

    def buscar_no_checadores(archivo_checadores):
        with open(archivo_checadores, 'r', newline='', encoding='utf-8') as file:
            csv_reader = csv.DictReader(file)
            data_checadores = list(csv_reader)

        # Crear un diccionario para almacenar las fechas de las checadas por PIN
        checadas_por_pin = defaultdict(list)
        for empleado in data_checadores:
            pin = empleado["PIN"]
            nombre = empleado["Nombre de empleado"]
            dispositivo = empleado["Dispositivo"]
            checada_str = empleado["Checada"]
            # Convertir la cadena de fecha y hora en un objeto datetime
            formato = determinar_formato(checada_str)
            checada_dt = datetime.strptime(checada_str, formato)
            # Almacenar la fecha de la checada en el diccionario
            checadas_por_pin[pin].append((checada_dt, nombre, dispositivo))
        # Obtener la fecha mínima global y máxima global de todas las checadas
        todas_las_fechas = [checada.date() for checadas in checadas_por_pin.values() for checada, _, _ in checadas]
        primer_dia_global = min(todas_las_fechas)
        ultimo_dia_global = max(todas_las_fechas)

        # Crear una lista con todas las fechas dentro del rango de fechas
        todas_las_fechas_posibles = [primer_dia_global + timedelta(days=i) for i in range((ultimo_dia_global - primer_dia_global).days + 2)]
        # Crear un diccionario para almacenar los días en que no se checó por PIN
        dias_no_checados_por_pin = defaultdict(list)

        for pin, checadas in checadas_por_pin.items():
            # Obtener el conjunto único de fechas (días) en que se checó para este PIN
            dias_checados = set(checada.date() for checada, _, _ in checadas)
            # Construir días no checados por empleado
            for fecha in todas_las_fechas_posibles:
                if fecha not in dias_checados:
                    # Verificar si el PIN ya tiene una entrada en el diccionario
                    if pin in dias_no_checados_por_pin:
                        # Si ya tiene una entrada, agregar la fecha no checada, nombre y dispositivo a las listas existentes
                        dias_no_checados_por_pin[pin][0].append(fecha)
                    else:
                        # Si no tiene una entrada, crear una nueva lista para el PIN
                        dias_no_checados_por_pin[pin] = ([fecha], [nombre], [dispositivo])

        return dias_no_checados_por_pin, data_checadores

    # Crear ventana de Tkinter para seleccionar el archivo
    root = Tk()
    root.withdraw()  # Ocultar la ventana principal

    # Ventana emergente para explicar el rango de fechas esperado
    messagebox.showinfo("Rango de fechas",
                        "Por favor, asegúrese de seleccionar un archivo CSV que contenga registros de checadas de hasta una semana atrás hasta la fecha actual.")

    # Solicitar al usuario que seleccione el archivo CSV
    archivo_checadores = filedialog.askopenfilename(title="Seleccione el archivo CSV", filetypes=[("Archivos CSV", "*.csv")])

    # Solicitar al usuario que seleccione el archivo CSV
    # archivo_checadores = filedialog.askopenfilename(title="Seleccione el archivo CSV", filetypes=[("Archivos CSV", "*.csv")])

    # Buscar días en que no se checó por empleado
    dias_no_checados_por_empleado, data_checadores = buscar_no_checadores(archivo_checadores)

    # Crear un nuevo archivo de Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    # Encabezados
    ws['A1'] = "PIN"
    ws['B1'] = "Nombre"
    ws['C1'] = "Dispositivo"
    # Escribir encabezados de fechas dinámicamente
    fechas = sorted(set(fecha for dias_no_checados, _, _ in dias_no_checados_por_empleado.values() for fecha in dias_no_checados))

    for idx, fecha in enumerate(fechas, start=1):
        ws.cell(row=1, column=3+idx).value = f"{fecha}"

    def buscar_nombre_por_pin(pin):
        nombre_empleado = None
        for empleado in data_checadores:
            if empleado['PIN'] == pin:
                nombre_empleado = empleado['Nombre de empleado']
                break  # Si encuentras el empleado, puedes salir del bucle
        return nombre_empleado

    def buscar_dispositivo_por_pin(pin):
        nombre_empleado = None
        for empleado in data_checadores:
            if empleado['PIN'] == pin:
                nombre_empleado = empleado['Dispositivo']
                break  # Si encuentras el empleado, puedes salir del bucle
        return nombre_empleado

    # Escribir datos en el archivo de Excel
    row = 2
    for pin, (dias_no_checados, nombres, dispositivos) in dias_no_checados_por_empleado.items():
        if dias_no_checados:  # Verificar si hay días no checados
            # print(f"PIN: {pin}, Dias no checados: {dias_no_checados}, Nombres: {nombres}, Dispositivos: {dispositivos}")

            # Escribir el PIN, nombre del empleado y dispositivo
            ws.cell(row=row, column=1).value = int(pin)  # Convertir PIN a entero
            ws.cell(row=row, column=2).value = buscar_nombre_por_pin(pin)  # Suponiendo que el nombre siempre estará en la primera posición
            ws.cell(row=row, column=3).value = buscar_dispositivo_por_pin(pin)  # Suponiendo que el dispositivo siempre estará en la primera posición

            # Escribir las fechas en las columnas correspondientes
            for idx, fecha in enumerate(fechas, start=1):
                if fecha in dias_no_checados:
                    ws.cell(row=row, column=3+idx).value = fecha.strftime("%d/%m/%Y")
                else:
                    ws.cell(row=row, column=3+idx).value = "A"
            row += 1

    ws.delete_cols(ws.max_column)

    # Guardar el archivo de Excel
    nombre_base = "resultadoFaltas"
    extension = "xlsx"

    def comparar_archivos(archivo_seleccionado, archivo_todos):
        # Cargar datos del archivo seleccionado por el usuario
        with open(archivo_seleccionado, 'r', newline='', encoding='utf-8') as file:
            csv_reader = csv.DictReader(file)
            data_seleccionados = list(csv_reader)

        # Cargar datos del archivo todos.csv
        with open(archivo_todos, 'r', newline='', encoding='utf-8') as file:
            csv_reader = csv.DictReader(file)
            data_todos = list(csv_reader)

        # Encontrar empleados faltantes en el archivo seleccionado
        pins_seleccionados = set(empleado["PIN"] for empleado in data_seleccionados)
        empleados_faltantes = [empleado for empleado in data_todos if empleado["PIN"] not in pins_seleccionados]

        return empleados_faltantes

    # Llamada a la función para comparar archivos
    empleados_faltantes = comparar_archivos(archivo_checadores, 'todos.csv')

    # Si hay empleados faltantes, agregarlos a una nueva hoja en el archivo Excel
    if empleados_faltantes:
        ws_faltantes = wb.create_sheet(title="EmpleadosNoExist")

        # Escribir encabezados en la nueva hoja
        ws_faltantes.append(["PIN", "Nombre"])

        # Escribir datos de empleados faltantes en la nueva hoja
        for empleado in empleados_faltantes:
            ws_faltantes.append([empleado["PIN"], empleado["Nombre"]])

    # nombre_archivo = f"{os.path.splitext(archivo_checadores)[0]}.xlsx"
    # wb.save(nombre_archivo)
    nombre_archivo = obtener_nombre_archivo(nombre_base, extension)
    wb.save(nombre_archivo)

    # # Indicar al usuario que se ha creado el archivo
    # print(f"Se ha creado el archivo '{nombre_archivo}' con éxito.")
    messagebox.showinfo("Archivo creado", f"Se ha creado el archivo '{nombre_archivo}' con éxito.")
except Exception as e:
    messagebox.showerror("Error", f"Se ha producido un error: {str(e)}")